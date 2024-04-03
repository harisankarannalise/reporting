import argparse
import copy
import json
import logging
import os
from datetime import datetime
from os import path
from openpyxl import Workbook
from openpyxl.styles import Alignment

from data_uploader import model_interface
from parse_fortis import generate_txt_report

output_location = "output/"
now = datetime.now()
log_file = path.join(output_location, f'get_{now.strftime("%Y%m%d")}.log')
logging.basicConfig(
    level=logging.DEBUG,
    filename=log_file,
    filemode='a',
    format='%(asctime)s %(name) -14s - %(levelname) -8s - %(message)s',
    datefmt='%H:%M:%S %d/%m/%y ',
)
logger = logging.getLogger(__name__)


def main(api_host, client_id, client_secret):
    # mi = model_interface.ModelInterface(
    #     api_host, client_id, client_secret, wait_time=5, max_workers=1
    # )
    #
    # results = mi.get_studies()
    # data = results.json()
    #
    # # Extract accession numbers
    # accessions = [study["accessionNumber"] for study in data["studies"]]
    # accessions_list_for_prediction = accessions[0:20]
    #
    # # Get prediction results from BE
    # results = mi.bulk_get(accessions_list_for_prediction)
    # for result in results:
    #     json_file = path.join(output_location, f"{result['accession']}.json")
    #     with open(json_file, 'w') as fp:
    #         json.dump(result, fp)
    #
    # # Generate txt report
    # generate_txt_report('output', 'reports')

    # Create excel report
    # Create a new Workbook
    wb = Workbook()

    # Select the active worksheet
    ws = wb.active

    # Define column headings
    column_headings = [
        "Accession Number",
        "List of Findings",
        "Report Embedded",
        "Link to Report Folder",
        "Link to Secondary Capture Folder"
    ]

    # Add column headings to the first row
    ws.append(column_headings)
    pwd = os.getcwd()

    accessions_list_for_prediction = ['GUBEW9QGY4ASYUPL', 'PUN5YZFXNJUHMMKT', 'K2PHME2TSFZKJPOG', 'FVMZKRJCHSF3BZ6U', 'PWZAOGPDS6F44GVM', 'GVNHRIX3USQYAXHY', 'HCR9WTFX9MRBKTHM', 'E2NCJWFWMZXXVFEE', 'MA8EID3SSFRQYPK6', 'IFY7KSWEYPYTZUHR', '4H2JXTHT9G9UIXFI', 'PT4QKF3QEEYCLNNF', '29b2328160a72d32', '2b08a0a785f6b', 'cc64e893acd99', '3dddd317e4b7e', 'd12c5785a86a', '87c1ff9de4abd', 'b9c43ecebd12f', '932062ca961fc']
    for accession_number in accessions_list_for_prediction:
        with open(f'reports/{accession_number}.txt', 'r') as file:
            # Read all lines from the file into a list
            lines = file.readlines()

        with open(f'output/{accession_number}.json', 'r') as json_file:
            model_output = json.load(json_file)

        finding_label = []

        relevant_findings = model_output["classification"]["findings"]["vision"]["study"]["classifications"]["relevant"]
        for group in relevant_findings:
            findings_list = group["findings"]
            for finding in findings_list:
                finding_label.append(f'{finding["labelName"]}\n')

        embedded_string = ''.join(lines)
        findings_string = ''.join(finding_label)

        folder_path = os.path.join(pwd, 'reports', f'{accession_number}.txt')
        link_to_folder = f'file://{folder_path}'

        new_row_data = [
            accession_number,  # Accession Number
            findings_string,  # List of Findings
            embedded_string,  # Report Embedded
            link_to_folder,  # Link to Report Folder
            "http://example.com/capture4"  # Link to Secondary Capture Folder
        ]
        ws.append(new_row_data)

        # Set the hyperlink for the "Link to Report Folder" cell
        cell = ws.cell(row=ws.max_row, column=4)
        cell.hyperlink = link_to_folder

        for col in range(1, len(new_row_data) + 1):
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='general',
            vertical='bottom',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0)

    # Save the workbook
    wb.save("accession_data.xlsx")

    print("Excel file created successfully with the specified columns using openpyxl.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "-c",
        "--client_id",
        help="Annalise Backend Client ID",
    )
    parser.add_argument(
        "-s",
        "--client_secret",
        help="Annalise Backend Client Secret",
    )
    parser.add_argument(
        "-a",
        "--api_host",
        help="Annalise Backend Client base url",
    )
    args = parser.parse_args()

    if (
        args.api_host is not None
        and args.client_id is not None
        and args.client_secret is not None
    ):
        main(
            args.api_host,
            args.client_id,
            args.client_secret
        )
    else:
        if args.input_file is not None and args.accession_number is not None:
            logger.error("Only specify an input file OR an accession number")
        else:
            logger.error("Arguments provided not sufficient to run. Use --help for more info.")
